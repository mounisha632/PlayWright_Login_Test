import asyncio
import os
from datetime import datetime

import pytest
from openpyxl import load_workbook
from playwright.async_api import async_playwright


# Get the directory where this script is located
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "accounts_playwright_unique_passwords.xlsx")

os.makedirs(os.path.join(SCRIPT_DIR, "logs"), exist_ok=True)

LOG_FILE = os.path.join(SCRIPT_DIR, "logs/register_results.txt")


def write_log(message):

    with open(LOG_FILE, "a", encoding="utf-8") as file:

        file.write(message + "\n")


async def register_user(
        page,
        gender,
        first_name,
        last_name,
        email,
        password):

    try:

        await page.goto(
            "https://demowebshop.tricentis.com/register"
        )

        await page.wait_for_load_state()

        # Gender

        if gender.lower() == "female":

            await page.get_by_label(
                "Female", exact=True
            ).check()

        else:

            await page.get_by_label(
                "Male", exact=True
            ).check()

        # First Name

        await page.get_by_label(
            "First name:"
        ).fill(first_name)

        # Last Name

        await page.get_by_label(
            "Last name:"
        ).fill(last_name)

        # Email

        await page.get_by_label(
            "Email:"
        ).fill(email)

        # Password

        await page.get_by_label(
            "Password:",
            exact=True
        ).fill(password)

        # Confirm Password

        await page.get_by_label(
            "Confirm password:"
        ).fill(password)

        # Register

        await page.get_by_role(
            "button",
            name="Register"
        ).click()

        # Wait for page

        await page.wait_for_load_state()

        logout = page.get_by_role(
            "link",
            name="Log out"
        )

        # Check registration success

        if await logout.is_visible():

            msg = (
                f"{datetime.now()} | "
                f"PASS | "
                f"{email}"
            )

            print(msg)

            write_log(msg)

            await logout.click()

        else:

            error = "Registration Failed"

            try:

                error = await page.locator(
                    ".validation-summary-errors"
                ).inner_text()

            except Exception:

                pass

            msg = (
                f"{datetime.now()} | "
                f"FAIL | "
                f"{email} | "
                f"{error}"
            )

            print(msg)

            write_log(msg)

    except Exception as e:

        msg = (
            f"{datetime.now()} | "
            f"EXCEPTION | "
            f"{email} | "
            f"{str(e)}"
        )

        print(msg)

        write_log(msg)


async def main():

    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")

    wb = load_workbook(EXCEL_FILE)

    ws = wb["Accounts"]

    async with async_playwright() as p:

        browser = await p.chromium.launch(
            headless=False
        )

        context = await browser.new_context()

        page = await context.new_page()

        for row in ws.iter_rows(
                min_row=2,
                values_only=True):

            if None in row:
                write_log(f"{datetime.now()} | SKIP | Row contains None values")
                continue

            gender, first_name, last_name, email, password = row

            await register_user(
                page,
                gender,
                first_name,
                last_name,
                email,
                password
            )

        await context.close()

        await browser.close()


@pytest.mark.asyncio
async def test_register_users():
    """Pytest test function for user registration"""
    await main()


if __name__ == "__main__":
    asyncio.run(main())