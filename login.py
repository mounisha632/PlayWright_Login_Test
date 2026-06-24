import asyncio
import os
import datetime
import pytest
from openpyxl import load_workbook
from playwright.async_api import async_playwright

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "accounts_playwright_unique_passwords.xlsx")
LOG_FILE = os.path.join(SCRIPT_DIR, "logs/login_results.txt")


def write_log(message):
    """Write message to log file"""
    with open(LOG_FILE, "a", encoding="utf-8") as file:
        file.write(message + "\n")


async def login_user(context, email, password):
    """Login a single user and log results"""
    # Create a new page for each login to avoid state issues
    page = await context.new_page()
    try:
        await page.goto("https://demowebshop.tricentis.com/")
        await asyncio.sleep(1)
        
        # Click on Log in link
        await page.get_by_role("link", name="Log in").click()
        await asyncio.sleep(1)
        
        # Fill email field using get_by_label like in demo code
        await page.get_by_label("Email:").click()
        await asyncio.sleep(0.3)
        await page.get_by_label("Email:").fill(email)
        await asyncio.sleep(0.5)
        
        # Fill password field using get_by_label like in demo code
        await page.get_by_label("Password:").click()
        await asyncio.sleep(0.3)
        await page.get_by_label("Password:").fill(password)
        await asyncio.sleep(0.5)
        
        # Click login button using exact selector from demo code
        await page.get_by_role("button", name="Log in").click()
        
        # Wait for home page - check if logout link appears or error appears
        logout_link = page.get_by_role("link", name="Log out")
        error_msg = page.locator(".validation-summary-errors")
        
        for _ in range(20): # wait up to 10 seconds (20 * 0.5s)
            if await logout_link.is_visible():
                write_log(f"{datetime.datetime.now()} | PASS | {email}")
                await logout_link.click()
                await asyncio.sleep(1)
                break
            elif await error_msg.is_visible():
                error_text = await error_msg.inner_text()
                error_text = error_text.replace('\n', ' ').strip()
                write_log(f"{datetime.datetime.now()} | FAIL | {email} | {error_text}")
                break
            await asyncio.sleep(0.5)
        else:
            write_log(f"{datetime.datetime.now()} | FAIL | {email} | Failed to reach home page")
            
    except Exception as e:
        write_log(f"{datetime.datetime.now()} | EXCEPTION | {email} | {str(e)}")
    finally:
        # Close the page after login attempt
        try:
            await page.close()
        except:
            pass


async def main():
    """Main function to orchestrate login tests"""
    
    # Validate Excel file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")
    
    # Load workbook and get active sheet
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    
    # Initialize log file
    write_log(f"\n{'='*80}")
    write_log(f"Login Test Started: {datetime.datetime.now()}")
    write_log(f"{'='*80}\n")
    
    # Launch browser
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=False)
        context = await browser.new_context()
        
        # Loop through Excel rows starting from row 2 (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows or rows with None values
            if not row or None in row:
                write_log(f"{datetime.datetime.now()} | SKIP | Row {row_num} contains None values")
                continue
            
            email, password = row[3], row[4]  # Email is column D (index 3), Password is column E (index 4)
            
            # Skip if email or password is empty
            if not email or not password:
                write_log(f"{datetime.datetime.now()} | SKIP | Row {row_num} has empty email or password")
                continue
            
            await login_user(context, email, password)
        
        # Close browser and context
        await context.close()
        await browser.close()
    
    write_log(f"\n{'='*80}")
    write_log(f"Login Test Completed: {datetime.datetime.now()}")
    write_log(f"{'='*80}\n")


@pytest.mark.asyncio
async def test_login_users():
    """Pytest test function for user login"""
    await main()


if __name__ == "__main__":
    asyncio.run(main())
